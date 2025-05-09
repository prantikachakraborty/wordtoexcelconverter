import openpyxl
from openpyxl.utils import get_column_letter

def duplicate_sheet(file_path, sheet_name):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, keep_vba=True)
        
        # Check if the sheet exists
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            return
        
        # Get the original sheet
        original_sheet = workbook[sheet_name]
        
        # Create a new sheet for the duplicate
        duplicate_sheet_name = f"{sheet_name}_copy"
        if duplicate_sheet_name in workbook.sheetnames:
            print(f"Sheet '{duplicate_sheet_name}' already exists. Overwriting it.")
            del workbook[duplicate_sheet_name]
        
        duplicate_sheet = workbook.create_sheet(title=duplicate_sheet_name)
        
        # Copy content from the original sheet to the duplicate
        for row in original_sheet.iter_rows():
            for cell in row:
                new_cell = duplicate_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                # Explicitly copy style attributes
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        
        # Copy column widths
        for col_idx, col_dim in original_sheet.column_dimensions.items():
            duplicate_sheet.column_dimensions[col_idx].width = col_dim.width
        
        # Copy row heights
        for row_idx, row_dim in original_sheet.row_dimensions.items():
            duplicate_sheet.row_dimensions[row_idx].height = row_dim.height
        
        # Save the workbook
        workbook.save(file_path)
        print(f"Duplicate sheet '{duplicate_sheet_name}' created successfully.")
    
    except Exception as e:
        print(f"An error occurred: {e}")

# Usage
file_path = "Sunflower_tailars_2024-25.xlsm"
sheet_name = "spouseguarantee"
duplicate_sheet(file_path, sheet_name)